import csv
import logging
import os
import sys
import time
from functools import partial

import matplotlib
import matplotlib.ticker as tkr
import numpy as np
import pandas as pd
# import statsmodels.api as sm
from matplotlib import pyplot as plt

from eam_core.common_analysis import convert, metric_conversion

logger = logging.getLogger(__name__)

from eam_core.graphical_analysis import get_platform_total, \
    plot_combined_platform_and_total, group_user_devices, plot_platform_process_annual_total, get_total_per_year, \
    get_total_monthly_average, group_by, filter_by, exclude_by, \
    plot_platform_process_per_device_hour
from eam_core.common_graphical_analysis import plot_kind, load_metadata, sum_interval
from eam_core.util import load_trace_data, kWh_p_J, load_df, load_as_df_qantity

# todo: this is all untested

def calculate_SRCs(input_variables, y):
    """
    Calculate standardised regression coefficients.
    See p. 9 ff. in "Sensitivity analysis in practice : a guide to assessing scientific models".
                     Andrea Saltelli et al. 2004

    For linear models, the SRCs (beta) are equal to $S^x_\sigma$ and the square of the SRCs sum up to 1:
    $1 = \sum (S_x^\sigma )^2 = \sum (\beta \frac{\sigma_x}{\sigma_y})^2$

    if the model coefficient of determination R^2_y is close to one (for monotone models) then the SRC^2 provide the percentage of output variance effected by the input variable

    :param input_variables:
    :param y:
    :return:
    """
    X = np.array(input_variables)
    X = sm.add_constant(X)
    ols_model = sm.OLS(y, X).fit()
    ols_model.summary()
    # compute standardised regression coefficients
    # compare with
    SRCs_square = np.array([b * v.std() / y.std() for v, b in zip(X[:, 1:].T, ols_model.params[1:])])
    R_squared = ols_model.rsquared

    return SRCs_square, R_squared


def get_total_footprint(process_footprint_dict, metric):
    y = sum(process_footprint_dict[metric].values())
    return y


def generate_plots_and_tables(scenario=None, metric='use_phase_energy', base_dir='.', start_date='2016-01-01',
                              end_date='2016-12-01', output_directory=None):
    metadata = load_metadata(output_directory, base_dir=base_dir)

    xlsx_file_name = f'{base_dir}/{output_directory}/results_{scenario}_{time.strftime("%m%d-%H%M")}.xlsx'
    writer = pd.ExcelWriter(xlsx_file_name)
    sheet_descriptions = {}
    pd.DataFrame.from_dict(sheet_descriptions, orient='index').to_excel(writer, 'toc')
    # df of x samples, monthly frequency between start and end date
    load_data = partial(load_as_df_qantity, f'{output_directory}/result_data.hdf5')

    # ======================== GO ================
    data = load_data()
    unit = str(data.units)
    # data.head(2).to_excel(writer, 'all data')
    sheet_descriptions['mean values'] = f'a direct load of the result data, monthly mean values. Unit: {unit}'
    logger.info("storing mean values to excel")
    data.m.mean(level='time').mean().to_excel(writer, 'mean values')

    # sum up monthly values to aggregate - duration depends on distance between start and end date
    #    load_data_aggegrate = lambda : load_data().sum(level='samples')
    xlabel = f'{metric_conversion[metric].split()[2]}/a'
    common_args = {'start_date': start_date,
                   'end_date': end_date,
                   'base_dir': base_dir,
                   'xlabel': xlabel,
                   'metadata': metadata,
                   'output_scenario_directory': output_directory}
    logger.info("plot_platform_process_annual_total")

    writer.save()
    sys.exit()

    plot_platform_process_annual_total(common_args, end_date, load_data, metadata, start_date, writer,
                                       platform='iPlayer', scenario=scenario)

    if scenario == 'baseline':
        plot_platform_process_annual_total(common_args, end_date, load_data, metadata, start_date, writer,
                                           platform='Terrestrial', scenario=scenario)

        plot_platform_process_annual_total(common_args, end_date, load_data, metadata, start_date, writer,
                                           platform='Satellite', scenario=scenario)
        plot_platform_process_annual_total(common_args, end_date, load_data, metadata, start_date, writer,
                                           platform='Cable', scenario=scenario)

    plot_platform_process_per_device_hour(common_args, end_date, load_data, metadata, start_date, writer,
                                          platform='iPlayer', scenario=scenario, base_dir=base_dir,
                                          output_directory=output_directory)

    logger.info("annual_platform_and_total")

    annual_platform_and_total(base_dir, common_args, end_date, load_data, metadata, scenario, sheet_descriptions,
                              start_date, writer, output_directory=output_directory)

    writer.save()
    sys.exit()

    # --------------------- plot_platform_total_by_device_hours ---------------
    plot_platform_total_by_device_hours(base_dir, common_args, end_date, load_data, metric, metadata, scenario,
                                        sheet_descriptions, start_date, writer, output_directory=output_directory)

    # --------------------- plot_all_processes_annual -------------------------
    data, unit = load_data()
    df = group_user_devices(data, metadata=metadata)
    df = sum_interval(df, start_date, end_date)
    d = plot_kind(df, figsize=(15, 12), file_name=f'{scenario}_plot_all_processes_annual.pdf', title='annual total',
                  **common_args)
    d.quantile([.25, .5, .75]).to_excel(writer, 'by device types')

    # --------- Various Groups Distribution and Non-distribution --------- --------- ---------
    logger.info("plot_all_processes_grouped")
    data, unit = load_data()
    # combine non-distribution processes

    df = group_by(data, metadata=metadata, group_name='TVs', included_types=['Viewing Device'], pattern_list=['TV'])
    df = group_by(df, metadata=metadata, group_name='Other Viewing Devices', included_types=['Viewing Device'],
                  pattern_list=['Smartphone', 'tablet', 'desktop', 'laptop', 'GamesConsole'])
    df = group_by(df, metadata=metadata, included_types=['STB', 'PVR'], group_name='STBs and PVRs')
    # df = group_by(df, metadata=metadata, included_types=['PVR'], group_name='PVRs')

    df = group_by(df, metadata=metadata, pattern_list=['WiFi Router'], group_name='WiFi Router')
    df = group_by(df, metadata=metadata, pattern_list=['Core', 'Access Network Port', 'CDN'],
                  group_name='IP Distribution')
    df = group_by(df, metadata=metadata, pattern_list=['Cable Central Infrastructure (IP)'],
                  group_name='Cable Distribtion')
    df = group_by(df, metadata=metadata, pattern_list=['Satellite Infrastructure'], group_name='Satellite Distribtion')
    df = group_by(df, metadata=metadata, pattern_list=['Terrestrial Infrastructure', 'DTT TV Amplififer'],
                  group_name='Terrestrial Distribtion')
    df = group_by(df, metadata=metadata, pattern_list=['Playout', 'CCM and Localisation Network', 'Video Factory'],
                  group_name='Playout')

    df = sum_interval(df, start_date, end_date)
    d = plot_kind(df, figsize=(15, 12), file_name=f'{scenario}_plot_all_processes_grouped.pdf',
                  title=f'Annual Total Energy Consumption [{unit}]',
                  **common_args)
    q = d.quantile([.25, .5, .75])
    q['sum'] = q.sum(axis=1)
    q.to_excel(writer, 'all_grouped')
    sheet_descriptions['all_grouped'] = f'All processes, arranged into groups. Unit: {unit}'

    # --------- Group Distribution and Non-distribution --------- --------- ---------
    data, unit = load_data()
    # combine non-distribution processes
    df = group_by(data, metadata=metadata, group_name='Non-Distribution',
                  included_types=home_equipment_devices_types)
    # the rest are distribution processes
    df = group_by(df, metadata=metadata, group_name='Distribution')
    df = sum_interval(df, start_date, end_date)
    d = plot_kind(df, figsize=(15, 12), file_name=f'{scenario}_plot_dist_non_dist_grouped.pdf', title='annual total',
                  **common_args)
    # --------- Non - Distribution Only --------- --------- ---------
    logger.info("plot non-dist processes")
    data, unit = load_data()
    # filter columns with device types in :field:home_equipment_devices_types
    df = filter_by(data, metadata=metadata, included_types=home_equipment_devices_types)
    df = sum_interval(df, start_date, end_date)
    d = plot_kind(df, figsize=(15, 12), file_name=f'{scenario}_plot_non_dist_processes.pdf', title='annual total',
                  **common_args)
    q = d.quantile([.25, .5, .75])
    q['sum'] = q.sum(axis=1)
    q.to_excel(writer, 'non_dist')
    sheet_descriptions['non_dist'] = f'Non distribution processes. Unit: {unit}'
    # --------- Distribution Only --------- --------- ---------
    data, unit = load_data()
    df = exclude_by(data, metadata=metadata, excluded_types=home_equipment_devices_types)
    df = sum_interval(df, start_date, end_date)
    d = plot_kind(df, figsize=(15, 12), file_name=f'{scenario}_plot_dist_processes.pdf', title='annual total',
                  **common_args)
    q = d.quantile([.25, .5, .75])
    q['sum'] = q.sum(axis=1)
    q.to_excel(writer, 'dist')
    sheet_descriptions['dist'] = f'Distribution processes. Unit: {unit}'

    # --------- STBs --------- --------- ---------
    data, unit = load_data()
    # filter columns with device types in :field:home_equipment_devices_types
    df = filter_by(data, metadata=metadata, included_types=['STB', 'PVR'])
    df = sum_interval(df, start_date, end_date)
    # d = plot_box(df, figsize=(15, 12), file_name=f'{scenario}_plot_non_dist_processes.pdf', title='annual total',
    #              **common_args)
    q = df.quantile([.25, .5, .75])
    q['sum'] = q.sum(axis=1)
    q.to_excel(writer, 'STBs')
    sheet_descriptions['STBs'] = f'STB and PVR processes. Unit: {unit}'

    # --------- STBs --------- --------- ---------
    data, unit = load_data()
    # filter columns with device types in :field:home_equipment_devices_types
    df = filter_by(data, metadata=metadata, included_types=['Viewing Device'], pattern_list=['TV'])
    df = sum_interval(df, start_date, end_date)
    # d = plot_box(df, figsize=(15, 12), file_name=f'{scenario}_plot_non_dist_processes.pdf', title='annual total',
    #              **common_args)
    q = df.quantile([.25, .5, .75])
    q['sum'] = q.sum(axis=1)
    q.to_excel(writer, 'TVs')
    sheet_descriptions['TVs'] = f'TV processes. Unit: {unit}'
    writer.save()

    logger.info("writing TOC")

    from openpyxl import load_workbook
    wb = load_workbook(xlsx_file_name)
    ws = wb['toc']
    # ----------------------------------------------------- TV vs STB
    # ------------------ TOC ----------------
    for row, (name, desc) in enumerate(sheet_descriptions.items(), start=1):
        ws.cell(coordinate=f'A{row}').value = name
        ws.cell(coordinate=f'B{row}').value = desc
        ws.cell(coordinate=f'C{row}').value = f'=HYPERLINK("#\'{name}\'!A1", "Link")'
    ws.column_dimensions["A"].width = '23'
    ws.column_dimensions["B"].width = '63'
    ws.column_dimensions["C"].width = '10'
    print(xlsx_file_name)
    wb.save(xlsx_file_name)


def annual_platform_and_total(base_dir, common_args, end_date, load_data, metadata, scenario, sheet_descriptions,
                              start_date, writer, output_directory=None):
    # plot relative to device hours
    device_secs_per_month_df = load_trace_data(output_directory, 'duration', base_dir=base_dir)
    device_hours_per_month_df = device_secs_per_month_df / 3600
    sheet_label = 'dev hrs per mth'
    device_hours_per_month_df.mean(level='time').to_excel(writer, sheet_label)
    sheet_descriptions[
        sheet_label] = f'a direct load of the device hours per month, mean values. Unit: hours per month'
    # ---------------------
    data, unit = load_data()
    series_total, sheet_description = get_total_per_year(data, metadata, False, device_hours_per_month_df,
                                                         start_date_=start_date, end_date_=end_date, unit=unit)
    sheet_label = f'tot per a'
    series_total.head(100).to_excel(writer, sheet_label)
    sheet_descriptions[sheet_label] = f'total footprint per year; over all processes. Unit: {unit}'
    # --------------------- combined_platform_and_total
    data, unit = load_data()
    data_by_platform = get_platform_total(data, device_hours_per_month_df, metadata)
    sheet_label = f'tot by platf'
    data_by_platform.mean(level='time').mean().to_excel(writer, sheet_label)
    sheet_descriptions[sheet_label] = f'total per platform per month. Unit: {unit}'
    data_by_platform = sum_interval(data_by_platform, start_date, end_date)
    data_by_platform.quantile([.25, .5, .75]).to_excel(writer, 'plat quantiles p a')
    series_total.rename('Total', inplace=True)
    # print(total.head(2))
    plot_combined_platform_and_total(series_total, data_by_platform, scenario,
                                     file_name=f'{scenario}_annual_platform_and_total.pdf', sharex=True,
                                     title=f'Annual Sum', **common_args, x_lable=f'{unit}',
                                     output_directory=output_directory, series_total_label=f'{series_total.name}')


def plot_platform_total_by_device_hours(base_dir, common_args, end_date, load_data, m, metadata, scenario,
                                        sheet_descriptions, start_date, writer, total=True, output_directory=None):
    if m == 'use_phase_energy':
        data, unit = load_df(scenario, m, base_dir=base_dir, output_directory=output_directory) * kWh_p_J, 'kWh'
    else:
        data, unit = load_data()
    device_secs_per_month_df = load_trace_data(output_directory, 'duration', base_dir=base_dir)
    device_hours_per_month_df = device_secs_per_month_df / 3600
    df_platform_data = get_platform_total(data, device_hours_per_month_df, metadata=metadata,
                                          relative_to_views=True)
    sheet_label = f'tot p plat p dev hrs'
    df_platform_data.mean(level='time').mean().to_excel(writer, sheet_label)
    sheet_descriptions[sheet_label] = f'total by platform by device hrs per month ({unit} per h)'
    # ---------------------
    if m == 'use_phase_energy':
        data, unit = load_df(scenario, m, base_dir=base_dir, output_directory=output_directory) * kWh_p_J, 'kWh'
    else:
        data, unit = load_data()
    series_total, sheet_description = get_total_monthly_average(data, metadata, True, device_hours_per_month_df,
                                                                start_date_=start_date, end_date_=end_date, unit=unit)
    sheet_label = 'tot per a per dev hr'
    series_total.head(10).to_excel(writer, sheet_label)
    series_total.rename('Total', inplace=True)
    sheet_descriptions[sheet_label] = sheet_description
    # -----------------------
    d, d_total = plot_combined_platform_and_total(series_total, df_platform_data, scenario,
                                                  xlabel=f'{unit}/device-hour',
                                                  file_name=f'{scenario}_plot_platform_total_by_device_hours.pdf',
                                                  device_seconds_per_month_df=device_hours_per_month_df,
                                                  relative_to_views=True,
                                                  title='Monthly total by platform relative to device hours',
                                                  formatter=tkr.FuncFormatter(lambda y, p: "{0:.2f}".format(y)),
                                                  sharex=True, output_directory=output_directory,
                                                  series_total_label=f'{series_total.name}',
                                                  **{k: v for k, v in common_args.items() if not k == 'xlabel'})
    sheet_label = 'platf quart p mth p hr'
    d.quantile([.25, .5, .75]).to_excel(writer, sheet_label)
    sheet_descriptions[sheet_label] = f'quantiles of total by platform by device hrs per month ({unit} per h)'


def sensitivity_analysis(process_footprint_dict, metric, model, output_directory=None):
    vars = model.collect_input_variables(filter=lambda x: type(x.val) == np.ndarray and x.val.size > 1,
                                         by_process=False, exclude_results=True)

    standard_vars = [var.val for var in vars.values()]
    d = np.asarray(standard_vars)

    y = get_total_footprint(process_footprint_dict, metric)

    # pickle results
    df = pd.DataFrame(data=d.T, columns=vars.keys())

    df[metric] = y
    df.to_pickle(f'{output_directory}/pd_pickles/input_variables_{metric}.pdpkl')

    srcs, r_squared = calculate_SRCs(d.T, y)

    with open(f'{output_directory}/{model.name}_results.csv', 'w') as f:
        writer = csv.writer(f)
        # print a table of SRCs
        writer.writerows(['', ''])
        writer.writerow(['SRCs'])
        writer.writerows([*zip(vars.keys(), srcs)])
        writer.writerows(['', ''])
        writer.writerow(['R^2', r_squared])

        try:
            df = pd.DataFrame.from_dict(process_footprint_dict[metric])
        except ValueError:
            df = pd.DataFrame.from_dict(process_footprint_dict[metric], orient='index')

        writer.writerows(['', ''])
        writer.writerow(['process results'])

        writer.writerow([df.mean()])
        flierprops = {'color': 'black', 'marker': 'x', 'alpha': 0.3}

        df.boxplot(rot=90, flierprops=flierprops)  # showfliers=False,
        matplotlib.pyplot.tight_layout()
        matplotlib.pyplot.savefig(f"{output_directory}/{model.name}_results_{metric}.png")
        writer.writerows(['', ''])
        writer.writerow(['Overall footprint'])
        writer.writerow([df.sum(axis=1).mean()])


def plot_results(name, metric, output_directory=None):
    if not os.path.exists(f'{output_directory}/plots/'):
        os.makedirs(f'{output_directory}/plots/')

    e, unit = convert(load_df(name, metric, output_directory=output_directory), metric)

    e = e.reindex_axis(e.mean().sort_values().index, axis=1)
    # @todo find out why one cannot set a distribution to 0 mean and 0 variance without result values to contain negative values - this should not happen...
    e = e.abs()
    e.mean(level='time').plot(kind='area', figsize=(10, 8))
    handles, labels = plt.gca().get_legend_handles_labels()
    legend = plt.legend(handles[::-1], labels[::-1], bbox_to_anchor=(0.95, 0.8), loc=2, \
                        borderaxespad=0., prop={'size': 8}, fancybox=True, framealpha=0.3)
    plt.title(f'{metric} [{unit}]')
    plt.savefig(f'{output_directory}/plots/area_over_time_{metric}.pdf')


home_equipment_devices_types = ['Network CPE', 'Amplifier', 'STB', 'PVR', 'Viewing Device']
