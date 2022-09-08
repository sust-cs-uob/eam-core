import os
import shlex
import subprocess

import matplotlib
import sys
import time

import numpy as np
from openpyxl.styles import Alignment

from eam_core import SimulationControl

matplotlib.use('Agg')
from matplotlib import pyplot as plt# , #MatplotlibDeprecationWarning

import json as json

from shutil import copyfile
from eam_core.util import get_sim_run_description, create_output_folder, \
    draw_graph_from_dotfile, load_as_df_quantity, load_as_plain_df, prepare_simulation,find_node_by_name

from eam_core.YamlLoader import YamlLoader
from eam_core.common_graphical_analysis import load_metadata, plot_kind, plot_process_with_input_vars

from eam_core.dataframe_result_utils import group_data

from eam_core.SimulationRunner import SimulationRunner

import pandas as pd
from ruamel import yaml

from functools import partial

# from joblib import Parallel, delayed
# import multiprocessing
import itertools
import configparser

import eam_core.log_configuration as logconf

logconf.config_logging()

# @todo: upgrade pandas 0.24 to stay compatible
import warnings

# warnings.filterwarnings("ignore", category=MatplotlibDeprecationWarning)

try:
    CONFIG_FILE = "local.cfg"
    cfg = configparser.ConfigParser()
    cfg.read_file(itertools.chain(['[global]'], open(CONFIG_FILE)))
    config = cfg['global']
except:
    config = {}

import logging

logger = logging.getLogger()


def setup_parser(args):
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--analysis_config', '-a', help="name of analysis_config to run")
    parser.add_argument('--comment', '-c',
                        help="Provide a comment to this simulation run. This will skip a UI prompt and can be used for headless environments.")
    parser.add_argument('--docker', '-d', help="generate graphviz graphs in a docker container", action='store_true')
    parser.add_argument('--save_pickle', '-sp', help="For grouped scenario-enabled runs, save a .pkl summary file of the run", action='store_true')
    parser.add_argument('--filetype', '-f', help="generate graphviz graphs in a docker container", default='pdf')
    parser.add_argument('--formula_checks', '-fc', help="perform checks on formulas and variables", action='store_true')
    parser.add_argument('--local', '-l', help="When using cloud datasources, don't check for updates cloud spreadsheets", action='store_true')
    parser.add_argument('--IDs', '-id', help="give each process and variable an ID", action='store_true')
    parser.add_argument('--multithreaded', '-m', help="use multicore speed ups", action='store_true')

    parser.add_argument('--verbose', '-v', help="enable debug level logging", action='store_true')
    parser.add_argument('yamlfile', help="yaml file to run")
    parser.add_argument('--sensitivity', '-n', help="run sensitivity analysis", action='store_true')
    parser.add_argument('--skip_documentation', '-sd', help="create documentation", action='store_true')
    parser.add_argument('--skip_storage', '-ss', help="do not store results", action='store_true')

    args = parser.parse_args(args)
    # print(args)
    return args


def write_formula(outstream, formula, name):
    outstream.write(f'## Process {name}\n\n')
    outstream.writelines(formula)
    outstream.write('\n\n')


def create_documentation(runner):
    """
    Generate markdown file with model code.
    :param model:
    :return:
    """
    from networkx.drawing.nx_pydot import to_pydot
    import networkx as nx
    import pypandoc
    model = runner.model
    output_directory = runner.sim_control.output_directory
    H = nx.relabel_nodes(model.process_graph, lambda n: n.name)
    # write_dot(H, 'models/graphs/baseline.dot')
    pydot = to_pydot(H)

    logger.debug("Building model documentation markdown")

    with open(f'{output_directory}/{model.name}_model_documentation.md', 'w') as f:
        f.write("# Model\n\n")

        for node in pydot.get_node_list():
            name = node.get_name().strip('"')
            process_node = find_node_by_name(model, name)

            method = process_node.formulaModel.formula.text
            write_formula(f, method, name)

            items = []

            logger.debug(f"Processing process {name}")
            # for item in input_vars[name].items():
            #     logger.debug(f"Processing variable {item[0]}")
            #     dataa = item[1]
            #     if isinstance(dataa, Q_):
            #         dataa = dataa.m
            #         metadata[item[0]] = {'unit': dataa.units}
            #     if dataa.index.nlevels == 2:
            #         item__mean = dataa.mean(level='time').mean()
            #     else:
            #         item__mean = dataa.mean()
            #     items.append((item[0], item__mean))
            #
            # # collect all the traces of this process
            # process_traces = {key: traces[key][name] for key in traces.keys() if name in traces[key]}
            #
            # for item in process_traces.items():
            #     if item[1].index.nlevels == 2:
            #         item__mean = item[1].mean(level='time').mean()
            #     else:
            #         item__mean = item[1].mean()
            #     items.append((item[0], item__mean))
            #
            # f.writelines(tabulate(items, ['variable', 'value'], tablefmt='simple'))

            f.write('\n\n')
            #
            # items = []
            # for metric, unit in zip(['use_phase_energy', 'use_phase_carbon', 'embodied_carbon'],
            #                         ['J', 'gCO2e', 'gCO2e']):
            #     if name + '_' + metric in df.columns:
            #         v = df[name + '_' + metric].mean()
            #         items.append([metric, v, unit])
            #
            # f.writelines(tabulate(items, ['variable', 'value', 'unit'], tablefmt='simple'))

            f.write('\n\n')

    if runner.use_docker:
        logger.info("writing pandoc")
        cwd = os.getcwd()

        # cmd = f"docker run -v {cwd}:{project_dir} -w {project_dir} markfletcher/graphviz dot {dot_file} -T{file_type} -Gsplines=ortho -Grankdir=LR -Gnodesep=0.1 -Gratio=compress"
        # cmd = f"docker run -v {cwd}:{project_dir} -w {project_dir} markfletcher/graphviz dot {dot_file} -T{file_type} -Gsplines=ortho -Grankdir=BT"
        # cmd = f"docker run -v {cwd}:{project_dir} -w {project_dir} markfletcher/graphviz dot {dot_file} -T{file_type} -Gsplines=ortho -Grankdir=BT > {dot_render_filename}"
        pdf_file_name = f"{runner.sim_control.output_directory}/{model.name}_model_documentation.pdf"

        cmd = f"docker run -v  $(pwd):/source jagregory/pandoc -f markdown -t latex {runner.sim_control.output_directory}/{model.name}_model_documentation.md -o {pdf_file_name} -V 'geometry:margin=0.2in, landscape'"
        logger.info(f'running docker cmd {cmd}')
        l_cmd = shlex.split(cmd)
        logger.info(f'running docker cmd {l_cmd}')
        # with open(pdf_file_name, 'w') as output:
        #     with subprocess.Popen(l_cmd, stdout=output) as proc:
        #         pass
        ps = subprocess.Popen(cmd, shell=True)
    else:
        # todo: untested without docker
        logger.info("writing pandoc")
        output = pypandoc.convert_file(f'{output_directory}/{model.name}_model_documentation.md', 'pdf',
                                       outputfile=f'{output_directory}/{model.name}_model_documentation.pdf',
                                       extra_args=['-V', 'geometry:margin=0.2in, landscape'])


def run_scenario(scenario, model_run_base_directory=None, simulation_run_description=None, yaml_struct=None, args=None,
                 analysis_config=None, output_persistence_config=None):
    """
    Run a scenario
    :param scenario:
    :type scenario:
    :param model_run_base_directory:
    :type model_run_base_directory:
    :param simulation_run_description:
    :type simulation_run_description:
    :param yaml_struct:
    :type yaml_struct:
    :param args:
    :type args:
    :param analysis_config:
    :type analysis_config:
    :param output_persistence_config:
    :type output_persistence_config:
    :param mean_run:
    :type mean_run:
    :return:
    :rtype:
    """
    logger.info(f'Evaluating scenario {scenario}')
    # prepare a sub directory
    model_output_directory = model_run_base_directory + f"/{scenario}"
    if not os.path.exists(model_output_directory):
        os.makedirs(model_output_directory)

    mean_run = None

    if yaml_struct['Metadata'].get('sample_mean', True) == False:
        # run with just mean values, so that we can deal with 'sub-zero' values during analysis
        # this is needed because of sampling effect when uncertainty is very high
        mean_run = run_mean(args, model_run_base_directory, simulation_run_description, yaml_struct, scenario)

    create_model_func, sim_control, yaml_struct = prepare_simulation(model_output_directory,
                                                                     simulation_run_description, yaml_struct,
                                                                     scenario, filename=args.yamlfile, IDs=args.IDs,
                                                                     formula_checks=args.formula_checks, args=args,
                                                                     use_time_series=True)
    if args.sensitivity:
        # todo: untested with args.sensitivity
        runner = SimulationRunner()
        # model, variances = runner.run_SA(create_model_func=create_model_func, embodied=False, sim_control=sim_control)
        model, correlations = runner.run_OTA_SA(create_model_func=create_model_func, embodied=False, sim_control=None)

        df = pd.DataFrame(correlations).T

        logger.debug(df)

        writer = pd.ExcelWriter(f'{model_output_directory}/sa.xlsx')
        df.to_excel(writer, 'OTA SA')
        writer.close()
        return (scenario, runner)

    runner = SimulationRunner()
    runner.use_docker = args.docker
    runner.sim_control = sim_control

    runner.run(create_model_func=create_model_func, debug=True,
               target_units=YamlLoader.get_target_units(yaml_struct),
               result_variables=yaml_struct['Analysis'].get('result_variables', []),
               output_persistence_config=output_persistence_config)

    analysis(runner, yaml_struct, analysis_config=analysis_config, mean_run=mean_run, image_filetype=args.filetype)
    # analysis model results

    if not args.skip_documentation:
        create_documentation(runner)
    logger.info(f'Evaluating scenario {scenario} complete')
    return (scenario, runner)


def run_mean(args, model_run_base_directory=None, simulation_run_description=None, yaml_struct=None,
             scenario='default'):
    """

    :param args:
    :type args:
    :return:
    :rtype:
    """

    logger.info(f"Running mean for scenario {scenario}")
    if not model_run_base_directory:
        model_run_base_directory, simulation_run_description, yaml_struct = load_configuration(args)

    import copy
    _yaml_struct = copy.deepcopy(yaml_struct)

    _yaml_struct['Metadata']['sample_mean'] = True
    _yaml_struct['Metadata']['sample_size'] = 1

    run_scenario_f = partial(run_scenario, model_run_base_directory=model_run_base_directory,
                             simulation_run_description=simulation_run_description, yaml_struct=_yaml_struct, args=args,
                             analysis_config={}, output_persistence_config={})

    _, runner = run_scenario_f(scenario)

    logger.info(f"Running mean for scenario {scenario} finished")
    return runner


def run(args, analysis_config=None):
    """

    :param args:
    :type args:
    :param analysis_config: if None, the analysis_config in the args is read
    :type analysis_config:
    :return:
    :rtype:
    """
    # initialise this set of simulations runs by creating a directory for output files
    model_run_base_directory, simulation_run_description, yaml_struct = load_configuration(args)

    # setup any analysis configuration
    if analysis_config is None:
        if args.analysis_config:
            # find the analysis_config in the yaml file by name
            analysis_configs_ = [item for item in yaml_struct['Metadata'].get('analysis_configs', []) if
                                 item['name'] == args.analysis_config]
            if analysis_configs_:
                logger.info(f"Running with analysis_config '{analysis_configs_[0]}'")
                analysis_config = analysis_configs_[0]
            else:
                logger.warning(f"Could not find analysis_config {args.analysis_config}")
                analysis_config = {}

        else:
            analysis_config = {}

    # scenarios to evaluate
    scenarios = yaml_struct['Analysis'].get('scenarios', [])
    if not 'default' in scenarios:
        # 'default' is implicit
        scenarios.append('default')
    # define aspects/data to store

    logger.info(f'Skip store results set to {args.skip_storage}')
    output_persistence_config = {'store_traces': not args.skip_storage,
                                 'process_traces': analysis_config.get('process_traces', [])}

    # a partial to invoke for each scenario
    run_scenario_f = partial(run_scenario, model_run_base_directory=model_run_base_directory,
                             simulation_run_description=simulation_run_description, yaml_struct=yaml_struct, args=args,
                             output_persistence_config=output_persistence_config, analysis_config=analysis_config)

    if args.multithreaded:
        # todo: this doesn't work, needs review
        logger.info("Running in parallel")
        from pathos.multiprocessing import ThreadPool
        import multiprocessing
        num_cores = multiprocessing.cpu_count()
        results = ThreadPool(processes=num_cores).map(run_scenario_f, scenarios)
        # results = Parallel(n_jobs=num_cores)(delayed(run_scenario_f)(i) for i in scenarios)
    else:
        results = []
        for scenario in scenarios:
            results.append(run_scenario_f(scenario))

    # combine all run results
    runners = dict((x, y) for x, y in results)

    analysis_config.update(yaml_struct['Analysis'])
    analysis_config.update(yaml_struct['Metadata'])

    if not args.sensitivity:
        scenario_paths = {scenario_name: run_data.sim_control.output_directory for scenario_name, run_data in
                          runners.items()}
        if args.analysis_config:
            summary_analysis(scenario_paths, model_run_base_directory, analysis_config, yaml_struct,
                             image_filetype=args.filetype, save_pickle=args.save_pickle)
    return runners


def load_configuration(args):
    """
    model_run_base_directory

    :param args:
    :type args:
    :return:
    :rtype:
    """
    if not 'comment' in args.__dict__:
        simulation_run_description = get_sim_run_description()
    else:
        simulation_run_description = args.comment
    yamlfile = args.yamlfile
    yaml_struct = None
    with open(yamlfile, encoding="utf-8", mode="r", errors='ignore') as stream:
        try:
            yaml_struct = yaml.load(stream, Loader=yaml.RoundTripLoader)
        except yaml.YAMLError as exc:
            logger.error(f'Error while loading yaml file {yamlfile} {exc}')
            sys.exit(1)
    model_basedir = f"output/{yaml_struct['Metadata']['model_name']}/"
    model_run_base_directory = create_output_folder(model_basedir)
    for file_location in yaml_struct['Metadata'].get('file_locations', []):
        # todo: this isn't tested. is it ever used?

        if file_location['type'] == 'google_spreadsheet':
            google_id_alias = file_location['google_id_alias']

            sheet_id_alias_ = config['google_drive_params_sheet_id_alias']
            sheet_id_alias = json.loads(sheet_id_alias_)
            file_id = sheet_id_alias[google_id_alias]

            if not args.local:
                from eam_core.gdrive_data_helper import update_local_data
                logger.info(f'updating cloud file with alias "{google_id_alias}"')
                from eam_core.gdrive_data_helper import update_local_data
                modDTime, revision = update_local_data(file_id, file_location['file_name'])

            file_name = os.path.basename(file_location['file_name'])
            copyfile(file_location['file_name'], f'{model_run_base_directory}/{file_name}')

    else:
        logger.info("Running with local parameter data only.")
    return model_run_base_directory, simulation_run_description, yaml_struct


def plot_scenario_comparison(scenario_paths, model_run_base_directory, base_dir, yaml_struct, image_filetype=None):
    load_data = load_as_df_quantity
    variable = yaml_struct['Metadata'].get('comparison_variable', 'energy')
    fig = plt.figure(figsize=(10, 5))
    ax = fig.add_subplot(111)

    numeric_values = {}
    max_ = 0
    for scenario_name, scenario_path in scenario_paths.items():
        output_directory = scenario_path
        # load dataframe of energy values
        pint_pandas_data, m = load_data(f'{output_directory}/result_data_{variable}.hdf5')
        units = {v[0]: v[1] for v in pint_pandas_data.pint.dequantify().columns.values}
        df = pint_pandas_data.pint.dequantify()
        df.columns = df.columns.droplevel(1)
        data = df
        base_dir = '.'
        metadata = load_metadata(output_directory, base_dir=base_dir)

        # if 'groups' in plot_def:
        #     data = group_data(data, metadata, plot_def)

        mean_ = data.mean(level='time').sum(axis=1)
        numeric_values[scenario_name] = mean_
        mean_.plot(ax=ax, kind='line',
                   # legend=False,
                   linewidth=1,
                   alpha=.3,
                   label=scenario_name
                   )
        mean__max = mean_.max()
        if mean__max > max_:
            max_ = mean__max

    ax.set_ylim(bottom=0, top=max_ * 1.1)
    ax.legend()

    plt.tight_layout()
    file_name = f'scenarios.{image_filetype}'
    if file_name:
        if not os.path.exists(f'{base_dir}/{model_run_base_directory}'):
            os.makedirs(f'{base_dir}/{model_run_base_directory}')
        file_name_ = f'{base_dir}/{model_run_base_directory}/{file_name}'
        logger.info(f'storing plot at {file_name_}')
        fig.savefig(file_name_)

    return numeric_values


def summary_analysis(scenario_paths, model_run_base_directory,
                     analysis_config, yaml_struct, image_filetype=None, save_pickle=False):
    """
    Compare all scenarios with each other.

    """
    base_dir = '.'

    xlsx_file_name = f'{base_dir}/{model_run_base_directory}/summary_{time.strftime("%m%d-%H%M")}.xlsx'
    writer = pd.ExcelWriter(xlsx_file_name)
    sheet_descriptions = {}
    pd.DataFrame.from_dict(sheet_descriptions, orient='index').to_excel(writer, 'toc')

    numeric_values = plot_scenario_comparison(scenario_paths, model_run_base_directory, base_dir, yaml_struct,
                                              image_filetype=image_filetype)

    for scenario, numeric_vals in numeric_values.items():
        numeric_vals.to_excel(writer, f'mean {scenario} (comp)')

    # ======================== GO ================
    variables = analysis_config.get('numerical', [])
    for variable in variables:
        data = None
        unit = None

        logger.info(f'numerical analysis for var {variable}')
        for scenario_name, scenario_path in scenario_paths.items():
            scenario = scenario_name
            logger.info(f'numerical analysis for scenario {scenario}')

            scenario_directory = model_run_base_directory + f'/{scenario}'

            if data is None:
                data, units = load_as_plain_df(f'{scenario_directory}/result_data_{variable}.hdf5')

                # data.rename(lambda x: f'{x}.{scenario}', axis='columns', inplace=True)

                if not yaml_struct['Metadata'].get('with_group', False):
                    logger.info(f'building excel dataframe for non-grouped {scenario} scenario')
                    data = data.mean(level='time').mean().to_frame(name=scenario)
                else:
                    logger.info(f'building excel dataframe for grouped {scenario} scenario')
                    data = data.reorder_levels(['group', 'time', 'samples']).groupby(level=['group']).mean()

                    # add an index level annotating output scenario
                    # this is needed as grouped scenarios are concatenated and need a way of separating!
                    data = pd.concat({scenario: data}, names=['Scenario'])

                # check all units are the same
                units_set = set(units[p] for p in units.keys())
                assert len(units_set) == 1

                unit = units_set.pop()

            else:
                if not yaml_struct['Metadata'].get('with_group', False):
                    logger.info(f'joining excel dataframe for non-grouped {scenario} scenario')
                    _dframe, _ = load_as_plain_df(f'{scenario_directory}/result_data_{variable}.hdf5')
                    # _dframe.rename(lambda x: f'{x}.{scenario}', axis='columns', inplace=True)

                    # data = pd.concat([s1, s2], axis=1)
                    data = data.join(_dframe.mean(level='time').mean().to_frame(name=scenario))
                else:
                    logger.info(f'concatenating excel dataframe for grouped {scenario} scenario')
                    data_scenario, _ = load_as_plain_df(f'{scenario_directory}/result_data_{variable}.hdf5')
                    data_scenario = data_scenario.reorder_levels(['group', 'time', 'samples']).groupby(level=['group']).mean()
                    data_scenario = pd.concat({scenario: data_scenario}, names=['Scenario'])
                    data = pd.concat([data, data_scenario])
                    if save_pickle:
                        logger.info(f'saving pickled dataframe for grouped {scenario} scenario')
                        data.to_pickle(model_run_base_directory + f'/result_data_full_{variable}.pkl')

        sheet_name = f'mean {variable}'
        sheet_descriptions[
            sheet_name] = f'{sheet_name}: a direct load of the result data, monthly mean values. Unit: {unit}'

        logger.info("storing mean values to excel")

        data.to_excel(writer, sheet_name)

    writer.save()
    write_TOC_to_excel(sheet_descriptions, variables, xlsx_file_name)


def write_TOC_to_excel(sheet_descriptions, variables, xlsx_file_name):
    logger.info("writing TOC")
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_file_name)
    ws = wb['toc']
    # ----------------------------------------------------- TV vs STB
    # ------------------ TOC ----------------
    for row, (name, desc) in enumerate(sheet_descriptions.items(), start=1):
        ws[f'A{row}'].value = desc
        ws[f'B{row}'].value = f'=HYPERLINK("#\'{name}\'!A1", "Link")'
    ws.column_dimensions["A"].width = '75'
    ws.column_dimensions["B"].width = '10'

    for ws_name in [sheet for sheet in wb.sheetnames if not sheet == 'toc']:
        logger.debug(f'applying column styles to {ws_name}')
        ws = wb[ws_name]
        style_result_worksheet(ws)
    wb.save(xlsx_file_name)


def style_result_worksheet(ws):
    ws.column_dimensions["A"].width = '60'
    ws.column_dimensions["B"].width = '14'

    for cell in ws['A']:
        cell.alignment = Alignment(horizontal='left')


def provide_dummy_for_missing_process_vals(data, metadata):
    """
    VERY HACKY function for providing data for processes that don't calculate energy/carbon
    We need data in the frame for it to be written in the results sheet, but we can't have strings directly
    So, fill in with dummy data that is written to results, then overwritten with strings later on.
    I'm using -1 for these dummy vals since they work with mean and *shouldn't* occur naturally (negative energy?)
    :param data: The dataframe with processes and the result for a variable. This is mutated!
    :param metadata: Stores information about processes, used to find missing vals.
    :return: Boolean of whether any dummy values have been added to the dataframe
    """
    return False # skip this while overwrite_dummy_vals_with_null is unimplemented.
    dummy_provided = False
    for process in metadata.keys():
        if not process in data:
            dummy_vals = np.full((data.values.shape[0]), -1)
            logger.info(f"storing dummy values for excel in {process}")
            data[process] = dummy_vals
            dummy_provided = True
    return dummy_provided


def overwrite_dummy_vals_with_null(sheets):
    """
    Given a list of sheets, iterate through them, find cells with the value -1 (indicating a dummy value),
    and replace it with 'null;, indicating a process has not calculated that variable.
    :param sheets: A list of the sheet names that need parsing.
    :return:
    """
    for sheet_name in sheets:
        # unimplemented.
        pass


def analysis(runner, yaml_struct, analysis_config=None, mean_run=None, image_filetype=None):
    """
    Performs analysis on a run within a given scenario.
    Creates any graphs, and creates the results_[scenario]_[datetime].xlsx sheet.

    :param runner:
    :param yaml_struct:
    :param analysis_config:
    :param mean_run:
    :param image_filetype:
    :return:
    """
    if analysis_config is None or not analysis_config:
        logger.info(f'analysis_config not provided, nothing to analyse')
        return
    model = runner.model
    sim_control = runner.sim_control
    scenario = sim_control.scenario

    # print(footprint_result_dict['use_phase_energy'].keys())
    # for k, v in footprint_result_dict['use_phase_energy'].items():
    #     print(k, v)
    # mean = SimulationRunner.get_stddev_and_mean(footprint_result_dict['use_phase_energy'], np.ones(2))
    # print(mean)

    if sim_control.use_time_series:
        output_directory = sim_control.output_directory

        # logger.info('generate_model_definition_markdown')
        # logger.warning('skipping markdown generation for now')
        # # @todo fix generate_model_definition_markdown
        # generate_model_definition_markdown(model, in_docker=args.docker, output_directory=output_directory)

        if 'start_date' in yaml_struct['Metadata']:
            start_date = yaml_struct['Metadata']['start_date']
            end_date = yaml_struct['Metadata']['end_date']

        if scenario == 'default':

            if 'process_tree' in analysis_config:
                process_group_colour_defs = analysis_config['process_tree'].get('process_group_colours', {})
                show_variables = analysis_config['process_tree'].get('show_variables', True)
                histograms = analysis_config['process_tree'].get('histograms', True)

                draw_graph_from_dotfile(model, show_variables=show_variables,
                                        file_type=image_filetype, metric='energy',
                                        start_date=start_date, end_date=end_date, colour_def=process_group_colour_defs,
                                        in_docker=runner.use_docker,
                                        output_directory=output_directory,
                                        target_units=YamlLoader.get_target_units(yaml_struct),
                                        edge_labels=show_variables,
                                        show_histograms=histograms
                                        )
        logger.info('generating plots and tables')
        # generate_plots_and_tables(scenario=scenario, metric='use_phase_energy', base_dir='.', start_date=start_date,
        #                           end_date=end_date, output_directory=output_directory)
        base_dir = '.'
        metadata = load_metadata(output_directory, base_dir=base_dir)

        xlsx_file_name = f'{base_dir}/{output_directory}/results_{scenario}_{time.strftime("%m%d-%H%M")}.xlsx'
        writer = pd.ExcelWriter(xlsx_file_name)
        sheet_descriptions = {}
        pd.DataFrame.from_dict(sheet_descriptions, orient='index').to_excel(writer, 'toc')
        # df of x samples, monthly frequency between start and end date
        load_data = load_as_df_quantity

        # ======================== GO ================
        variables = yaml_struct['Analysis'].get('numerical', [])
        for variable in variables:

            logger.info(f'numerical analysis for var {variable}')

            pint_pandas_data, m = load_data(f'{output_directory}/result_data_{variable}.hdf5')
            units = {v[0]: v[1] for v in pint_pandas_data.pint.dequantify().columns.values}
            df = pint_pandas_data.pint.dequantify()
            # df.columns = df.columns.droplevel(1)
            data = df

            dummy_provided = provide_dummy_for_missing_process_vals(data, metadata)

            unit = list(units.values())[0]

            sheet_name = f'total {variable} '
            sheet_descriptions[
                sheet_name] = f'{sheet_name}: total over assessment period. Unit: {unit}'
            logger.info("storing total values to excel")
            if sim_control.with_group:
                total = data.reorder_levels([2, 0, 1]).sort_index(level=['group', 'time']).mean(level=[0, 1]).sum(
                    level=0).T
            else:
                total = data.mean(level=0).sum()
            total.to_excel(writer, sheet_name)

            sheet_name = f'month mean {variable} '
            sheet_descriptions[
                sheet_name] = f'{sheet_name}: a direct load of the result data, monthly mean values. Unit: {unit}'
            logger.info("storing mean values to excel")
            mean = data.mean(level='time').mean()
            mean.to_excel(writer, sheet_name)

            use_f_units = yaml_struct['Analysis'].get('functional_units', False)
            if use_f_units:
                logger.info(f'Calculating functional units for {variable}...')
                data_per_f_unit, f_unit_type = extract_functional_units(yaml_struct, data, sim_control, variable)
                if data_per_f_unit is not None:
                    sheet_name = f'functional unit {variable}'
                    sheet_descriptions[
                        sheet_name] = f'{sheet_name}: The data presented in terms of {variable} per {f_unit_type}.'

                    mean_data_per_f_unit = data_per_f_unit.mean(level='time').mean()
                    mean_data_per_f_unit.to_excel(writer, sheet_name)
            else:
                logger.info('Functional units are not calculated for this model')

            if sim_control.sample_size > 1:
                # print(data)
                sheet_name = f'25 quantiles {variable}'
                sheet_descriptions[
                sheet_name] = f'{sheet_name}: a direct load of the result data, monthly mean values. Unit: {unit}'
                logger.info("storing quantile values to excel")
                low = data.abs().groupby(level=['time']).quantile(.25)
                low.to_excel(writer, sheet_name)

                sheet_name = f'75 quantiles {variable}'
                sheet_descriptions[
                sheet_name] = f'{sheet_name}: a direct load of the result data, monthly mean values. Unit: {unit}'
                data.abs().groupby(level=['time']).quantile(.75).to_excel(writer, sheet_name)

            if dummy_provided:
                overwrite_dummy_vals_with_null([f'total {variable} ', f'month mean {variable} ',
                                                f'25 quantiles {variable}', f'75 quantiles {variable}'])

        logger.info("plot_platform_process_annual_total")

        plot_defs = yaml_struct['Analysis'].get('plots', [])

        for plot_def in plot_defs:
            name = plot_def['name']
            if name in analysis_config.get('named_plots', []):
                plot_scenario_def(yaml_struct, plot_def, output_directory, start_date, end_date, base_dir,
                                  metadata, mean_run, name, scenario, image_filetype, load_data)
            else:
                logger.debug(f"Named plot {plot_def['name']} not used in analysis config")

        if 'individual_process_graphs' in analysis_config:
            # @todo - hack
            variable = yaml_struct['Metadata']['individual_process_graphs_variable']

            pint_pandas_data, m = load_data(f'{output_directory}/result_data_{variable}.hdf5')

            units = {v[0]: v[1] for v in pint_pandas_data.pint.dequantify().columns.values}
            # check all units are the same
            units_set = set(units[p] for p in units.keys())
            assert len(units_set) == 1

            unit = units_set.pop()

            df = pint_pandas_data.pint.dequantify()
            df.columns = df.columns.droplevel(1)
            data = df

            iv_dfs = plot_process_with_input_vars(model, sim_control, data, f'{scenario}_input_vars', output_directory,
                                                  unit, base_dir=base_dir, analysis_config=analysis_config,
                                                  image_filetype=image_filetype)

            # store in excel
            for proc, input_vars_df in iv_dfs.items():
                sheet_name = f'(iv) {proc}'[:30]
                sheet_descriptions[
                    sheet_name] = f'{sheet_name}: the input variables'
                input_vars_df.mean(level='time').to_excel(writer, sheet_name)

        if 'input_vars' in analysis_config:
            plot_input_vars(model, analysis_config, image_filetype, sim_control, base_dir, output_directory)

        logger.info("writing TOC")
        writer.save()
        write_TOC_to_excel(sheet_descriptions, variables, xlsx_file_name)

        return


def extract_functional_units(yaml_struct, data,  sim_control, variable):
    """
    Take the results for the specified output variable and calculate the results in terms of functional units.
    Functional units are specified in the Analysis section of the yaml model.
    todo needs tests!
    """
    f_unit_type = yaml_struct['Analysis'].get('functional_unit_type', "UNDEFINED")
    f_unit_var_list = yaml_struct['Analysis'].get('functional_unit_vars', [])

    f_unit_cumulative_value = 0
    parameter_set = sim_control.param_repo.parameter_sets
    scenario = sim_control.scenario
    for f_unit_var in f_unit_var_list:
        logger.info(f'Fetching functional unit value for {f_unit_var}')
        f_unit_value = parameter_set[f_unit_var].scenarios[scenario].kwargs.get('ref value', 0)
        f_unit_cumulative_value += f_unit_value
    logger.info(f'Total summed functional units: {f_unit_cumulative_value} {f_unit_type}')

    if f_unit_cumulative_value > 0:
        data_per_f_unit = data / f_unit_cumulative_value
        data_per_f_unit.columns = data_per_f_unit.columns.droplevel(1)
        #data_per_f_unit = data_per_f_unit.pint.quantify()
        return data_per_f_unit, f_unit_type
    else:
        logger.warning(f'{f_unit_type} value was not greater than 0. Value was {f_unit_cumulative_value}')
        return None, None


def plot_scenario_def(yaml_struct, plot_def, output_directory, start_date, end_date, base_dir,
                      metadata, mean_run, name, scenario, image_filetype, load_data):
    variable = plot_def['variable']
    logger.info(plot_def['name'])
    title = plot_def.get('title', 'Annual Total Energy Consumption')
    # sum up monthly values to aggregate - duration depends on distance between start and end date
    #    load_data_aggegrate = lambda : load_data().sum(level='samples')

    pint_pandas_data, m = load_data(f'{output_directory}/result_data_{variable}.hdf5')

    units = {v[0]: v[1] for v in pint_pandas_data.pint.dequantify().columns.values}
    df = pint_pandas_data.pint.dequantify()
    df.columns = df.columns.droplevel(1)
    data = df

    # check all units are the same
    units_set = set(units[p] for p in units.keys())
    assert len(units_set) == 1

    unit = units_set.pop()

    # sum up monthly values to aggregate - duration depends on distance between start and end date
    #    load_data_aggegrate = lambda : load_data().sum(level='samples')
    xlabel = f'{unit}/a'
    common_args = {'start_date': start_date,
                   'end_date': end_date,
                   'base_dir': base_dir,
                   'metadata': metadata,
                   'unit': unit,
                   'output_scenario_directory': output_directory}

    if 'xlabel' in plot_def:
        xlabel = plot_def['xlabel'].format(unit=unit)

    ylabel = None
    if 'ylabel' in plot_def:
        ylabel = plot_def['ylabel'].format(unit=unit)

    common_args.update({'xlabel': xlabel, 'ylabel': ylabel})

    kind = plot_def.get('kind', 'box')
    if kind == 'area':
        # for area charts if mean_run parameter is not null, we can assume this analysis here is of a run
        # with random sampling. In this case, we need the results from a 'non-sampling' run so zero-values
        # don't mess up the statistics
        mean_run_df = data
        if mean_run:
            mean_run_df, mean_run_metadata = mean_run.get_process_variable_values(variable,
                                                                                  YamlLoader.get_target_units(
                                                                                      yaml_struct),
                                                                                  mean_run.sim_control)

            mean_run_df = mean_run_df.pint.dequantify()
            mean_run_df.columns = mean_run_df.columns.droplevel(1)
    if 'groups' in plot_def:
        if mean_run:
            mean_run_df = group_data(mean_run_df, metadata, plot_def)
            common_args.update({'mean_data': mean_run_df})

        data = group_data(data, metadata, plot_def)

    data.to_pickle(f'{output_directory}/graph_data_{name}.pdpkl')

    d = plot_kind(data, figsize=(15, 12), file_name=f'{scenario}_{name}_{variable}.{image_filetype}',
                  title=title,
                  kind=kind, **common_args)


def plot_input_vars(model, analysis_config, image_filetype, sim_control, base_dir, output_directory):
    logger.info("plotting input vars")
    logger.debug("collecting input vars from model")
    all_vars = model.collect_input_variables()

    _vars = {}

    iv_ac = None

    for _, p_vars in all_vars.items():
        _vars.update(p_vars)

    # remove all vars that were not defined in the analysis config
    input_vars_ac = analysis_config['input_vars']
    if input_vars_ac is not None and 'variables' in input_vars_ac:
        iv_ac = set(input_vars_ac['variables'])
        _vars = {k: v for k, v in _vars.items() if k in iv_ac}
    logger.debug(f"plotting vars: {_vars}")

    file_name = f'input_vars.{image_filetype}'

    import copy
    rcParams_bkp = copy.deepcopy(plt.rcParams)
    plt.rcParams['axes.grid'] = True
    plt.rcParams['axes.linewidth'] = .2
    plt.rcParams['axes.spines.top'] = True
    plt.rcParams['axes.spines.right'] = True
    plt.rcParams['axes.spines.bottom'] = True
    plt.rcParams['axes.spines.left'] = True
    plt.rcParams['font.size'] = 8
    plt.rcParams['grid.linestyle'] = ':'
    plt.rcParams['lines.linewidth'] = .1
    plt.rcParams['hatch.linewidth'] = 0.3

    import math
    n = len(_vars)
    # n = 50
    rows = math.ceil(n / 3)
    f, axarr = plt.subplots(rows, 3, sharex='col', sharey=False)
    f.set_size_inches(15, n * 0.5)

    var_names = sorted(_vars.keys())
    for i, ax in enumerate(axarr.flat):
        if i == len(var_names):
            break
        var_name = var_names[i]
        val = _vars[var_name]
        logger.info(f"plotting input variable {var_name}")
        # for var_name, val in sorted(_vars.items()):
        #     fig = plt.figure(figsize=(10, 5))
        #     ax = fig.add_subplot(111)

        try:
            data = _vars[var_name].pint.to_reduced_units()
        except:
            logger.error(f"could not reduce units {_vars[var_name].pint.units}")
            pass

        p_df = pd.DataFrame(data=data.pint.m)
        p_df.columns = [var_name]
        if not isinstance(p_df.index, pd.MultiIndex):
            p_df.index = sim_control._df_multi_index

        mean_ = p_df.mean(level='time')
        mean_.plot(ax=ax, kind='line',
                   legend=False,
                   linewidth=1,
                   color='k', alpha=.3,
                   #                    marker="x"
                   )

        ax.set_ylabel(data.pint.units)
        #     fig.title(var_name)
        ax.set_title(var_name[:70], fontsize=7, loc='right')
        ax.yaxis.get_offset_text().set_size(6)
        ax.yaxis.get_offset_text().set_y(0)
        # https://jakevdp.github.io/PythonDataScienceHandbook/04.10-customizing-ticks.html
        ax.xaxis.set_major_locator(plt.MaxNLocator(3))

    # f.suptitle(var_name, fontsize=16)
    plt.subplots_adjust(hspace=0.4, wspace=.3)

    # f.tight_layout()
    if file_name:
        if not os.path.exists(f'{base_dir}/{output_directory}'):
            os.makedirs(f'{base_dir}/{output_directory}')
        file_name_ = f'{base_dir}/{output_directory}/{file_name}'
        logger.info(f'storing plot at {file_name_}')
        f.savefig(file_name_)
    plt.close('all')
    plt.rcParams.update(rcParams_bkp)


if __name__ == '__main__':
    args = setup_parser(sys.argv[1:])
    logger.info(f"Running with parameters {args}")
    if args.verbose:
        level = logging.DEBUG
        logger = logging.getLogger()
        logger.setLevel(level)
        for handler in logger.handlers:
            handler.setLevel(level)

    runners = run(args)
